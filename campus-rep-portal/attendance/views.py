from rest_framework import generics, permissions
from .models import LectureSession
from rest_framework import serializers
from .serializers import LectureSessionSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse, FileResponse
from django.utils.text import get_valid_filename
from django.db import models
from rest_framework import status
from .models import LectureSession, AttendanceRecord, Notification, CampusDocument
from .utils import haversine_distance
from .models import Announcement
from .serializers import AnnouncementSerializer, CampusDocumentSerializer, CampusDocumentUploadSerializer
from rest_framework.exceptions import ValidationError
from .models import AuditLog
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from accounts.models import User
from accounts.push import send_email_to_user, send_push_to_user

class AnnouncementListView(generics.ListAPIView):
    serializer_class = AnnouncementSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Announcement.objects.filter(
            department=self.request.user.department,
            level=self.request.user.level
        )


class CheckInView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        if not request.user.matric_number:
            return Response(
                {"detail": "Please add your matric number before checking in."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            session = LectureSession.objects.get(pk=pk, is_active=True)
        except LectureSession.DoesNotExist:
            return Response({"detail": "No active session found."}, status=status.HTTP_404_NOT_FOUND)

        student_lat = request.data.get('latitude')
        student_lon = request.data.get('longitude')

        if student_lat is None or student_lon is None:
            return Response({"detail": "latitude and longitude are required."}, status=status.HTTP_400_BAD_REQUEST)

        distance = haversine_distance(
            float(student_lat), float(student_lon),
            session.latitude, session.longitude
        )

        if distance > session.radius_meters:
            return Response(
                {"detail": f"Too far from venue. You are {int(distance)}m away, limit is {session.radius_meters}m."},
                status=status.HTTP_403_FORBIDDEN
            )

        record, created = AttendanceRecord.objects.get_or_create(student=request.user, session=session)

        if not created:
            return Response({"detail": "Already checked in."}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"detail": "Checked in successfully.", "distance_meters": int(distance)}, status=status.HTTP_201_CREATED)

class IsClassRep(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == 'CLASS_REP'


class LectureSessionCreateView(generics.CreateAPIView):
    queryset = LectureSession.objects.all()
    serializer_class = LectureSessionSerializer
    permission_classes = [IsClassRep]

    def perform_create(self, serializer):
        session = serializer.save(
            department=self.request.user.department,
            is_active=False,
            has_ended=False,
        )
        AuditLog.objects.create(
            rep=self.request.user,
            action='CREATED',
            course_code=session.course_code,
            venue_name=session.venue_name,
        )
        students = User.objects.filter(
            department=session.department, level=session.level, role='STUDENT'
        )
        for student in students.iterator():
            Notification.objects.create(
                user=student, type='SESSION',
                title=f'{session.course_code} session created',
                body=f'{session.venue_name} · {session.level} Level. Your class rep has created a new attendance session.',
            )
            if not student.session_notifications:
                continue
            try:
                send_push_to_user(
                    student,
                    f'{session.course_code} session created',
                    f'New session at {session.venue_name}. Check CampusPulse for details.',
                )
                send_email_to_user(
                    student, f'{session.course_code} session created',
                    f'Your class rep created a new {session.course_code} session at {session.venue_name}. Open CampusPulse for details.',
                )
            except Exception:
                pass
class LectureSessionToggleView(APIView):
    permission_classes = [IsClassRep]

    def post(self, request, pk):
        try:
            session = LectureSession.objects.get(
                pk=pk,
                department=request.user.department,
            )
        except LectureSession.DoesNotExist:
            return Response({'detail': 'Session not found.'}, status=status.HTTP_404_NOT_FOUND)

        if session.has_ended:
            return Response(
                {'detail': 'This session has already ended and cannot be reopened. Create a new session instead.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if session.is_active:
            session.is_active = False
            session.has_ended = True
            session.save(update_fields=['is_active', 'has_ended'])
            AuditLog.objects.create(
                rep=request.user,
                action='ENDED',
                course_code=session.course_code,
                venue_name=session.venue_name,
            )
            return Response({'detail': 'Session ended.', 'session': LectureSessionSerializer(session).data})

        session.is_active = True
        session.save(update_fields=['is_active'])
        AuditLog.objects.create(
            rep=request.user,
            action='STARTED',
            course_code=session.course_code,
            venue_name=session.venue_name,
        )
        students = User.objects.filter(
            department=session.department,
            level=session.level,
            role='STUDENT',
        )
        for student in students.iterator():
            Notification.objects.create(
                user=student, type='SESSION',
                title=f'{session.course_code} is now live',
                body=f'Attendance is open at {session.venue_name}.',
            )
            if not student.session_notifications:
                continue
            try:
                send_push_to_user(
                    student,
                    f'{session.course_code} is now live',
                    f'Check in at {session.venue_name} — attendance is open now.',
                )
                send_email_to_user(
                    student, f'{session.course_code} is now live',
                    f'Attendance is now open at {session.venue_name}. Check in through CampusPulse.',
                )
            except Exception as e:
                print(f"PUSH SEND FAILED: {e}")
        return Response({'detail': 'Session started.', 'session': LectureSessionSerializer(session).data})

class AnnouncementCreateView(generics.CreateAPIView):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    permission_classes = [IsClassRep]

    def perform_create(self, serializer):
        announcement = serializer.save(
            department=self.request.user.department,
            level=self.request.user.level,
            posted_by=self.request.user,
        )
        students = User.objects.filter(
            department=announcement.department, level=announcement.level, role='STUDENT'
        )
        for student in students.iterator():
            Notification.objects.create(
                user=student, type='ANNOUNCEMENT',
                title=announcement.title, body=announcement.body,
            )
            if not student.announcement_notifications:
                continue
            try:
                send_push_to_user(student, announcement.title, announcement.body[:180])
                send_email_to_user(student, announcement.title, announcement.body)
            except Exception:
                pass

class ActiveSessionsView(generics.ListAPIView):
    serializer_class = LectureSessionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return LectureSession.objects.filter(
            department=self.request.user.department,
            level=self.request.user.level,
            is_active=True
        )
class MySessionsView(generics.ListAPIView):
    serializer_class = LectureSessionSerializer
    permission_classes = [IsClassRep]

    def get_queryset(self):
        return LectureSession.objects.filter(
            department=self.request.user.department
        ).order_by('-created_at')
class MyStatsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    ELIGIBILITY_THRESHOLD = 70

    def get(self, request):
        eligible_sessions = LectureSession.objects.filter(
            department=request.user.department,
            level=request.user.level
        ).order_by('-created_at')

        total_sessions = eligible_sessions.count()
        attended = AttendanceRecord.objects.filter(
            student=request.user,
            session__in=eligible_sessions
        ).count()
        rate = round((attended / total_sessions) * 100) if total_sessions > 0 else 0

        streak = 0
        for session in eligible_sessions:
            if AttendanceRecord.objects.filter(student=request.user, session=session).exists():
                streak += 1
            else:
                break

        eligibility_status = "eligible" if rate >= self.ELIGIBILITY_THRESHOLD else "at_risk"
        classes_until_risk = None

        if eligibility_status == "eligible" and total_sessions > 0:
            n = 0
            while True:
                n += 1
                projected_rate = (attended / (total_sessions + n)) * 100
                if projected_rate < self.ELIGIBILITY_THRESHOLD:
                    classes_until_risk = n - 1
                    break
                if n > 100:
                    classes_until_risk = 100
                    break

        return Response({
            "attended": attended,
            "total_sessions": total_sessions,
            "rate": rate,
            "streak": streak,
            "eligibility_status": eligibility_status,
            "classes_until_risk": classes_until_risk,
            "eligibility_threshold": self.ELIGIBILITY_THRESHOLD,
        })
class LectureSessionDeleteView(generics.DestroyAPIView):
    serializer_class = LectureSessionSerializer
    permission_classes = [IsClassRep]

    def get_queryset(self):
        return LectureSession.objects.filter(department=self.request.user.department)

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            rep=self.request.user,
            action='DELETED',
            course_code=instance.course_code,
            venue_name=instance.venue_name,
        )
        instance.delete()

class ExportAttendanceView(APIView):
    permission_classes = [IsClassRep]

    def get(self, request, pk):
        try:
            session = LectureSession.objects.get(pk=pk, department=request.user.department)
        except LectureSession.DoesNotExist:
            return Response({"detail": "Session not found."}, status=status.HTTP_404_NOT_FOUND)

        records = AttendanceRecord.objects.filter(session=session).select_related('student')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        bold = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center')

        ws.merge_cells('A1:B1')
        ws['A1'] = f"{session.course_code} — Attendance Sheet"
        ws['A1'].font = title_font

        ws['A2'] = f"Venue: {session.venue_name}"
        ws['A3'] = f"Date: {session.created_at.strftime('%B %d, %Y')}"
        ws['A2'].font = Font(size=10)
        ws['A3'].font = Font(size=10)

        header_row = 5
        ws.cell(row=header_row, column=1, value="S/N").font = bold
        ws.cell(row=header_row, column=2, value="Full Name").font = bold
        ws.cell(row=header_row, column=3, value="Matric Number").font = bold

        for col in range(1, 4):
            ws.cell(row=header_row, column=col).border = thin_border
            ws.cell(row=header_row, column=col).alignment = center

        row = header_row + 1
        for i, record in enumerate(records, start=1):
            full_name = record.student.get_full_name() or record.student.username
            matric = record.student.matric_number or "N/A"

            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=full_name).border = thin_border
            ws.cell(row=row, column=3, value=matric).border = thin_border
            row += 1

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

        sig_row = row + 2
        ws.cell(row=sig_row, column=1, value="Class Rep's Signature: _______________________").font = Font(size=10)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{session.course_code}_{session.venue_name}_attendance.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response


class IsStudentOrRep(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in {'STUDENT', 'CLASS_REP'}


class CampusDocumentListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        return CampusDocumentUploadSerializer if self.request.method == 'POST' else CampusDocumentSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = CampusDocument.objects.select_related('department', 'uploaded_by')
        if user.role != 'SUPER_ADMIN':
            if not user.department:
                return queryset.none()
            # Keep academic documents inside the user's department and level.
            queryset = queryset.filter(department=user.department, level=user.level)
        query = (self.request.query_params.get('q') or '').strip()
        if query:
            queryset = queryset.filter(
                models.Q(title__icontains=query)
                | models.Q(description__icontains=query)
                | models.Q(course_code__icontains=query)
                | models.Q(uploaded_by__first_name__icontains=query)
                | models.Q(uploaded_by__last_name__icontains=query)
            )
        return queryset

    def perform_create(self, serializer):
        if self.request.user.role not in {'STUDENT', 'CLASS_REP'}:
            raise ValidationError('Only students and class representatives can upload documents.')
        if not self.request.user.department or not self.request.user.level:
            raise ValidationError('Complete your department and level before uploading a document.')
        level = serializer.validated_data.get('level')
        if level != self.request.user.level:
            raise ValidationError({'level': 'You can only upload documents for your current level.'})
        serializer.save(
            department=self.request.user.department,
            uploaded_by=self.request.user,
        )


class CampusDocumentDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            document = CampusDocument.objects.select_related('department').get(pk=pk)
        except CampusDocument.DoesNotExist:
            return Response({'detail': 'Document not found.'}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role != 'SUPER_ADMIN' and (
            document.department_id != getattr(request.user.department, 'id', None)
            or document.level != request.user.level
        ):
            return Response({'detail': 'You do not have access to this document.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            document.file.open('rb')
        except (FileNotFoundError, OSError, ValueError):
            return Response({'detail': 'The document file is unavailable.'}, status=status.HTTP_404_NOT_FOUND)

        filename = get_valid_filename(document.file.name.rsplit('/', 1)[-1]) or f'document-{document.pk}.pdf'
        response = FileResponse(document.file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class NotificationSerializer(serializers.ModelSerializer):
    class Meta:
        model = Notification
        fields = ['id', 'type', 'title', 'body', 'is_read', 'created_at']


class AuditLogSerializer(serializers.ModelSerializer):
    rep_name = serializers.CharField(source='rep.get_full_name', read_only=True)

    class Meta:
        model = AuditLog
        fields = ['id', 'action', 'course_code', 'venue_name', 'rep_name', 'timestamp']


class AuditLogView(generics.ListAPIView):
    serializer_class = AuditLogSerializer
    permission_classes = [IsClassRep]

    def get_queryset(self):
        return AuditLog.objects.filter(rep__department=self.request.user.department)

class NotificationListView(generics.ListAPIView):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)[:50]


class NotificationReadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            notification = Notification.objects.get(pk=pk, user=request.user)
        except Notification.DoesNotExist:
            return Response({'detail': 'Notification not found.'}, status=status.HTTP_404_NOT_FOUND)
        notification.is_read = True
        notification.save(update_fields=['is_read'])
        return Response({'detail': 'Notification marked as read.'})


class NotificationReadAllView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({'detail': 'All notifications marked as read.'})


from .models import ClassCode, generate_class_code


class MyClassCodeView(APIView):
    permission_classes = [IsClassRep]

    def get(self, request):
        code_obj, created = ClassCode.objects.get_or_create(
            department=request.user.department,
            level=request.user.level,
        )
        return Response({"code": code_obj.code})

    def post(self, request):
        code_obj, created = ClassCode.objects.get_or_create(
            department=request.user.department,
            level=request.user.level,
        )
        code_obj.code = generate_class_code()
        code_obj.save()
        return Response({"code": code_obj.code})
class MyHistoryView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        sessions = LectureSession.objects.filter(
            department=request.user.department,
            level=request.user.level,
        ).order_by('-created_at')

        attended_ids = set(
            AttendanceRecord.objects.filter(
                student=request.user, session__in=sessions
            ).values_list('session_id', flat=True)
        )

        data = [
            {
                'id': s.id,
                'course_code': s.course_code,
                'venue_name': s.venue_name,
                'date': s.created_at,
                'status': 'attended' if s.id in attended_ids else 'missed',
            }
            for s in sessions
        ]
        return Response(data)